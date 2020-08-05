from typing import Tuple

from openpyxl import Workbook

from app.exceptions import *


class ImportExcel():
    def __init__(self, file: object):
        if not self._check_file(file):
            raise BadRequestException
        self._file = file

    def _check_file(self, file) -> bool:
        return type(file) == Workbook

    def read_cells(self, *args: Tuple[str]) -> list:
        headers_list = [i.upper() for i in list(args)]
        values_in_file = []

        sheet = self._file.active

        cells = [i for i in sheet.iter_rows(sheet.min_row + 1, sheet.max_row, sheet.min_column, sheet.max_column, True)]
        head = [j for j in sheet.iter_rows(sheet.min_row, sheet.min_row, sheet.min_column, sheet.max_column, True)]

        for i in cells:
            dicio = {}
            for j in head:
                for k in range(len(i)):
                    key = j[k]
                    value = i[k]
                    if key.upper() in headers_list:
                        if type(value) == str:
                            value = value.capitalize()
                        dicio.update({key.lower(): value})
                values_in_file.append(dicio)
        return values_in_file
