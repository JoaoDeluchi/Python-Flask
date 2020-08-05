import tempfile
from typing import Tuple
import  os

from openpyxl import load_workbook, Workbook

from app.exceptions import BadRequestException


class ExportExcel():
    def __init__(self, file_name: str):
        self._test_dir = tempfile.mkdtemp()
        if not file_name:
            raise BadRequestException
        self.file_name = str(os.path.join(self._test_dir, file_name + '.xlsx'))
        self.book = self._verify_file()
        self.sheet = self.book.active

    def _verify_file(self) -> Workbook:
        try:
            return load_workbook(self.file_name)
        except:
            return Workbook()

    def save_table_sheet(self, *args: Tuple) -> str:
        if len(args) == 2:
            self.sheet[args[0]] = args[1]
            self.book.save(self.file_name)
            return self.file_name
        if len(args) == 3:
            self.sheet.cell(args[0],args[1],args[2])
            self.book.save(self.file_name)
            return self.file_name
        raise BadRequestException
