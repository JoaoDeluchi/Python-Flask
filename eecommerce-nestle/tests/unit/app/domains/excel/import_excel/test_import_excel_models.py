import tempfile
import unittest
from unittest.mock import patch

import openpyxl

from app.domains.excel.import_excel.models import ImportExcel
from app.exceptions import BadRequestException


class TestImportExcelModels(unittest.TestCase):

    def setUp(self) -> None:
        self._test_dir = tempfile.mkdtemp()

    def test_try_open_workbook_should_be_false(self):
        # Arrange
        # Act
        with self.assertRaises(BadRequestException) as ex:
            excel = ImportExcel('')
        # Assert
        self.assertEqual(str(ex.exception.description), 'Bad Request Exception')
        self.assertEqual(str(ex.exception.code), '400')

    # @patch.object(type.mro())
    @patch('openpyxl.workbook.workbook.Workbook.active')
    @patch.object(openpyxl.workbook.workbook.Worksheet,'iter_rows')
    @patch('app.domains.excel.import_excel.models.type')
    @patch.object(openpyxl,'load_workbook')
    def test_read_cells_should_print_12_times(self, load_workbook_mock, mock_type, iter_rows_mock, active_mock):
        #Arrange
        mock_type.side_effect = [openpyxl.Workbook,str,str,str,str,str,str]
        iter_rows_mock.side_effect = [(('peter','22','33332222'),('John','44','32322323'),('Steve','79','997625212')),[('nome','idade','fone')]]
        active_mock.iter_rows = iter_rows_mock
        load_workbook_mock.active = active_mock
        # Act
        excel = ImportExcel(load_workbook_mock)
        result = excel.read_cells('nome', 'idade')

        # Assert
        self.assertEqual(7,mock_type.call_count)
        self.assertEqual(2,iter_rows_mock.call_count)
        self.assertEqual(excel._file,load_workbook_mock)
        self.assertEqual(load_workbook_mock.active,active_mock)
        self.assertEqual(3, len(result))
