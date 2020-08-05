import io
import os
import tempfile
import unittest
from unittest.mock import patch, Mock, MagicMock

import openpyxl
from openpyxl import Workbook

from app.domains.category import AlreadyExistsException
from app.domains.category import CategoryDoNotExistException, CategoryInactiveException
from app.domains.category.actions import create, get, get_by_id, update, delete, upload_file, export_file,\
    get_providers_from_category
from app.domains.excel.import_excel.models import ImportExcel
from app.exceptions import BadRequestException


class TestCategoryActions(unittest.TestCase):

    def setUp(self) -> None:
        self.data = dict(name='Nan Supreme I')

    @patch('app.domains.category.actions.Category')
    @patch('app.domains.category.actions.save')
    def test_should_create_new_category(self, save_mock, category_mock):
        # Arrange

        # Action
        create(self.data)

        # Assert
        save_mock.assert_called_once()
        category_mock.assert_called_once()
        calls_category = category_mock.call_args.kwargs
        self.assertEqual(calls_category['name'], 'Nan Supreme I')

    @patch('app.domains.category.actions.Category')
    def test_should_return_all_categories(self, category_mock):
        # Arrange
        category_mock.id = '123'
        category_mock.name = 'Nan Supreme I'
        query = Mock()
        query.filter_by().all = MagicMock(return_value=[category_mock])
        category_mock.query = query

        # Action
        categories = get({})

        # Assert
        category_mock.query.filter_by().all.assert_called_once()
        self.assertEqual(len(categories), 1)

    @patch('app.domains.category.actions.Category')
    def test_should_return_category_by_id(self, category_mock):
        # Arrange
        category_mock.id = 'nestle'
        category_mock.name = 'Nestogeno Etapa 1'
        query = Mock()
        query.get = MagicMock(return_value=category_mock)
        category_mock.query = query

        # Action
        get_by_id('nestle')

        # Assert
        category_mock.query.get.assert_called_once_with('nestle')
        self.assertEqual(category_mock.name, 'Nestogeno Etapa 1')
        self.assertEqual(category_mock.id, 'nestle')

    @patch('app.domains.category.actions.commit')
    @patch('app.domains.category.actions.get_by_id')
    @patch('app.domains.category.actions.Category')
    def test_should_update_category(self, category_mock, get_by_id_mock, commit_mock):
        # Arrange
        category_mock.id = '123'
        category_mock.name = 'NESLAC'
        get_by_id_mock.return_value = category_mock

        # Action
        update('id', {'name': 'Updated'})

        # Assert
        commit_mock.assert_called_once()
        get_by_id_mock.assert_called_once()
        self.assertEqual(category_mock.name, 'Updated')

    @patch('app.domains.category.actions.commit')
    @patch('app.domains.category.actions.get_by_id')
    @patch('app.domains.category.actions.Category')
    def test_delete_should_delete_category(self, category_mock, get_by_id_mock, commit_mock):
        # Arrange
        category_mock.id = '12345'
        category_mock.name = 'Cafe'
        category_mock.is_active = True
        get_by_id_mock.return_value = category_mock

        # Action
        delete('12345')

        # Assert
        self.assertFalse(category_mock.is_active)
        commit_mock.assert_called_once()
        self.assertEqual(get_by_id_mock.call_count, 1)

    @patch.object(ImportExcel, 'read_cells')
    @patch('app.domains.excel.import_excel.models.type')
    @patch('app.domains.category.actions.load_workbook')
    @patch('app.domains.category.actions.Category')
    @patch('app.domains.category.actions.save')
    def test_should_import_categories_from_a_xlsx_file(self, save_mock, category_mock, load_workbook_mock, type_mock,
                                                       read_cells_mock):
        # Arrange
        file = {'name': 'categories.xlsx'}
        file['name'] = tempfile.mkdtemp() + '/categories.xlsx'
        file['file'] = (io.BytesIO(bytes(1)), 'categories.xlsx')
        load_workbook_mock.return_value = Workbook
        type_mock.return_value = Workbook
        read_cells_mock.return_value = [{'name': 'premium', 'profit_percent': '5%'}]
        # Action
        upload_file(file['name'])

        # Assert
        save_mock.assert_called_once()
        category_mock.assert_called_once()
        calls_category = category_mock.call_args.kwargs
        self.assertEqual('premium', calls_category['name'])
        self.assertEqual('5%', calls_category['profit_percent'])
        self.assertNotEqual('7%', calls_category['profit_percent'])

    @patch('app.domains.category.actions.get')
    @patch.object(ImportExcel, 'read_cells')
    @patch('app.domains.excel.import_excel.models.type')
    @patch('app.domains.category.actions.load_workbook')
    @patch('app.domains.category.actions.Category')
    @patch('app.domains.category.actions.save')
    def test_should_import_categories_with_repeated_values(self, save_mock, category_mock, load_workbook_mock, type_mock,
                                                       read_cells_mock,get_mock):
        # Arrange
        file = {'name': 'categories.xlsx'}
        file['name'] = tempfile.mkdtemp() + '/categories.xlsx'
        file['file'] = (io.BytesIO(bytes(1)), 'categories.xlsx')
        load_workbook_mock.return_value = Workbook
        type_mock.return_value = Workbook
        read_cells_mock.return_value = [{'name': 'premium', 'profit_percent': '5%'}]
        obj = Mock()
        obj.name = 'premium'
        get_mock.return_value = [obj]
        # Action
        upload_file(file['name'])

        # Assert
        get_mock.assert_called_once()
        self.assertEqual(0,category_mock.call_count)
        self.assertEqual(0,save_mock.call_count)

    def test_should_raise_exception_importing_xlsx_file(self):
        # Arrange
        file = {'name': 'categories.xlsx'}
        file['name'] = tempfile.mkdtemp() + '/categories.txt'
        file['file'] = (io.BytesIO(bytes(1)), 'categories.txt')
        # Act
        with self.assertRaises(BadRequestException) as ex:
            upload_file(file)
        # Assert
        self.assertEqual(str(ex.exception.description), 'Bad Request Exception')
        self.assertEqual(str(ex.exception.code), '400')

    @patch.object(tempfile, 'mkdtemp')
    @patch.object(openpyxl.workbook.workbook.Workbook, 'save')
    @patch('app.domains.category.actions.get')
    def test_should_export_categories_to_a_xlsx_file(self, get_category_mock, save_mock, temp_mock):
        # Arrange
        obj = Mock()
        obj.serialize.return_value = {'id': '123', 'name': 'premium', 'profit_percent': '5%'}
        get_category_mock.return_value = [obj]
        temp_mock.return_value = 'arquivo'

        # Action
        result = export_file()

        # Assert
        temp_mock.assert_called_once_with()
        get_category_mock.assert_called_once()
        self.assertEqual(result, os.path.join('arquivo', 'categories.xlsx'))

    @patch('app.domains.category.actions.get')
    def test_should_raise_exception_creating_category(self,mock_get):
        obj = Mock()
        obj.name = 'name'
        mock_get.return_value = [obj]
        with self.assertRaises(AlreadyExistsException) as ex:
            create({'name':'name'})
        self.assertEqual(str(ex.exception.description), 'Name already exists in Database')
        self.assertEqual(str(ex.exception.code), '409')
    @patch('app.domains.category.actions.Category')
    def test_should_raise_exception_when_try_to_get_by_id_an_inexistent_category(self, category_mock):
        # Arrange
        query = Mock()
        query.get = MagicMock(return_value=None)
        category_mock.query = query

        # Action
        with self.assertRaises(CategoryDoNotExistException) as ex:
            get_by_id('suco')

        # Assert
        self.assertEqual(str(ex.exception.code), '404')
        self.assertEqual(str(ex.exception.description), 'Category do not exist')

    @patch('app.domains.category.actions.get_by_id')
    @patch('app.domains.category.actions.Category')
    def test_should_raise_exception_when_try_to_update_an_inactive_category(self, category_mock, get_by_id_mock):
        # Arrange
        get_by_id_mock.return_value = category_mock
        category_mock.is_active = False

        # Action
        with self.assertRaises(CategoryInactiveException) as ex:
            update('', {})

        # Assert
        get_by_id_mock.assert_called_once()
        self.assertEqual(str(ex.exception.code), '403')
        self.assertEqual(str(ex.exception.description), 'Unable to update an inactive category')

    @patch('app.domains.category.actions.Category')
    def test_get_categories_from_provider_by_association(self, category_mock):
        # Arrange
        category_id = "152d3621-2b5d-450a-b751-d58e74d0c296"
        laticinios_provider_mock = MagicMock()
        laticinios_provider_mock.serialize.return_value = {
            'id': '4d459ea0-7171-45b2-8f44-c551a7c33593',
            'name': 'Laticinios'
        }

        texteis_provider_mock = MagicMock()
        texteis_provider_mock.serialize.return_value = {
            'id': '21146ed1-c6cf-4dda-85a8-ec26337fce18',
            'name': 'Têxteis'
        }

        providers_list = [laticinios_provider_mock, texteis_provider_mock]
        provider_categories_mock = MagicMock(provider=providers_list)
        category_mock.query.get = MagicMock()
        category_mock.query.get.return_value = provider_categories_mock

        # Action
        response = get_providers_from_category(category_id)

        # Assert
        self.assertIsInstance(response, list)
        self.assertEqual(response, [{'id': '4d459ea0-7171-45b2-8f44-c551a7c33593', 'name': 'Laticinios'},
                                    {'id': '21146ed1-c6cf-4dda-85a8-ec26337fce18', 'name': 'Têxteis'}])
        category_mock.query.get.assert_called_once_with(category_id)