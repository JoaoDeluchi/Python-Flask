import io
import os
import unittest
from unittest.mock import patch, Mock, MagicMock

import pytest
from openpyxl import Workbook

from app.domains.excel.import_excel.models import ImportExcel
from app.domains.products import ProductDoNotExistException, ProductInactiveException, NoDataToExportException
from app.domains.products.actions import create, get, get_by_id, update, delete, export_product, upload_file, \
    get_by_name
from app.domains.products.models import Product
import openpyxl
import tempfile

from app.domains.products.validate_product import ValidateCreateProduct
from app.exceptions import ProductExceptions


class TestProduct(unittest.TestCase):

    def setUp(self) -> None:
        self.data = {
            'name': 'NAN SUPREME 1',
            'cost_values': 22.49,
            'product_line': 'Crianças',
            'unit_per_box': 12,
            'weight_per_unit': 400,
            'shelf_life': 21,
            'sku': 12323196,
            'description': 'é muito bom'
        }

    @patch.object(ValidateCreateProduct, 'check_if_product_is_valid')
    @patch('app.domains.products.actions.save')
    def test_create_should_be_return_product(self, mock_save, validate_mock):
        data = {
            "name": "pão c maçã", "cost_values": 28.75,
            "unit_per_box": 6, "weight_per_unit": 1000,
            "measure_unit": "gramas", "shelf_life": "1234T",
            "sku": "45663567385687890", "description": "Suco muito bom",
            "id_product_line": "658f1406-1aca-4cd1-a872-f34f371202ee"
        }
        mock_save.return_value = data
        created_product = create(data)
        validate_mock.called_once()
        mock_save.assert_called_once()
        self.assertEqual(created_product, data)

    @patch.object(ValidateCreateProduct, 'check_if_product_is_valid')
    @patch('app.domains.products.actions.Product')
    @patch('app.domains.products.actions.save')
    def test_create_must_be_raise_an_exception(self, mock_save, mock_product, validate_mock):
        data = {
            "name": "pão c maçã", "cost_values": 28.75,
            "unit_per_box": 6, "weight_per_unit": 1000,
            "measure_unit": "gramas", "shelf_life": "1234T",
            "sku": "45663567385687890", "description": "Suco muito bom",
            "id_product_line": "658f1406-1aca-4cd1-a872-f34f371202ee"
        }
        mock_product.return_value = None
        validate_mock.return_value = ['msg']
        mock_save.side_effect = ProductExceptions(validate_mock)
        with pytest.raises(ProductExceptions) as exc:
            create(data)
        self.assertEqual(exc.value.error_list, ['msg'])
        self.assertEqual(exc.value.code, 406)

    @patch('app.domains.products.actions.Product')
    def test_get_should_be_return_all_products(self, products_mock):
        # Arrange
        product = Mock()
        product.name = 'Ninho'
        query = Mock()
        query.filter_by().all = MagicMock(return_value=[product])
        products_mock.query = query

        # Action
        products = get({})

        # Assert
        self.assertTrue(products_mock.query.filter_by().all.called_once())
        self.assertEqual(len(products), 1)

    @patch('app.domains.products.actions.Product')
    def test_get_by_name_should_be_return_an_product(self, mock_product):
        name = 'NameofProduct'
        product = Mock()
        product.name = name
        product.sku = '1234567890'
        query = Mock()
        query.filter().first = MagicMock(return_value=product)
        mock_product.query = query

        get_product = get_by_name(name)
        self.assertEqual(product, get_product)
        mock_product.called_once_with(name)

    @patch('app.domains.products.actions.Product')
    def test_get_by_id_should_be_return_product(self, products_mock):
        # Arrange
        product_test = Mock()
        product_test.id = '123'
        product_test.name = 'Ninho'
        query = Mock()
        query.get = MagicMock(return_value=product_test)
        products_mock.query = query

        # Action
        product = get_by_id('123')

        # Assert
        products_mock.query.get.assert_called_once_with('123')
        self.assertEqual(query.get.call_count, 1)
        self.assertEqual(product_test, product)

    @patch('app.domains.products.actions.Product')
    def test_get_by_id_must_return_an_exception(self, product_mock):
        _id = ' '
        query = Mock()
        query.get = MagicMock(return_value=ProductDoNotExistException)
        product_mock.query.get.return_value = None
        with pytest.raises(ProductDoNotExistException) as exc:
            get_by_id(_id)
        self.assertEqual(exc.value.code, 404)
        self.assertEqual(exc.value.description, 'Product do not exist in database')


    @patch('app.domains.products.actions.commit')
    @patch('app.domains.products.actions.get_by_id')
    @patch('app.domains.products.actions.Product')
    def test_update_products(self, products_mock, get_by_id_mock, commit_mock):
        # Arrange
        product_test = Mock()
        product_test.name = 'Ninho'
        product_test.cost_values = 210
        get_by_id_mock.return_value = product_test
        products_mock = product_test

        # Actions
        update('id', {'name': 'SuperNan',
                      'cost_values': 10})

        # Assert
        get_by_id_mock.assert_called_once()
        commit_mock.assert_called_once()
        self.assertEqual('SuperNan', products_mock.name)
        self.assertEqual(10, products_mock.cost_values)

    @patch('app.domains.products.actions.commit')
    @patch('app.domains.products.actions.get_by_id')
    def test_delete_products(self, get_by_id_mock, commit_mock):
        # Arrange
        products_mock = Product()
        _id = '49560967589'
        get_by_id_mock.return_value = MagicMock(is_active=True)

        # Actions
        delete(_id)

        # Assert
        get_by_id_mock.assert_called_once_with(_id)
        commit_mock.assert_called_once()
        self.assertFalse(products_mock.is_active)

    @patch('app.domains.products.actions.get_by_id_product_line')
    @patch.object(tempfile, 'mkdtemp')
    @patch.object(openpyxl.workbook.workbook.Workbook, 'save')
    @patch('app.domains.products.actions.get')
    def test_should_export_products_to_an_excel_file(self, get_product_mock, save_mock, temp_mock,
                                                     mock_get_by_id_product_line):
        product_mock = Mock()
        product_mock.serialize.return_value = {'id': '12312', 'name': 'nescau', 'cost_values': '12',
                                               'unit_per_box': '20',
                                               "weight_per_unit": "200g", "shelf_life": "24 meses", "sku": "158974362",
                                               "description": "Muito bom para tomar café da manha :D",
                                               'id_product_line': '2389'}
        get_product_mock.return_value = [product_mock]
        temp_mock.return_value = 'arquivo'
        product_line_mock = Mock()
        product_line_mock.name = 'name'
        mock_get_by_id_product_line.return_value = product_line_mock

        result = export_product()

        temp_mock.assert_called_once_with()
        get_product_mock.assert_called_once()
        self.assertEqual(result, os.path.join(temp_mock.return_value, 'products.xlsx'))

    @patch('app.domains.products.actions.rename_headers')
    def test_must_raise_exception_when_try_to_export_products(self, mock_rename_headers):
        mock_rename_headers.return_value = Exception
        with pytest.raises(NoDataToExportException) as exc:
            export_product()
        self.assertEqual(exc.value.code, 422)
        self.assertEqual(exc.value.description, 'No data to export')

    @patch('app.domains.products.actions.Product')
    @patch('app.domains.products.actions.get_by_id')
    def test_should_raise_exception_when_try_to_update_an_inactive_product(self, get_by_id_mock,
                                                                           products_mock):
        # Arrange
        get_by_id_mock.return_value = products_mock
        products_mock.is_active = False

        # Action
        with self.assertRaises(ProductInactiveException) as ex:
            update('', {})

        # Assert
        get_by_id_mock.assert_called_once()
        self.assertEqual(str(ex.exception.code), '403')
        self.assertEqual(str(ex.exception.description), 'Unable to update an inactive product')

    @patch.object(ValidateCreateProduct, 'check_if_product_of_excel_is_valid')
    @patch('app.domains.products.actions.Product')
    @patch('app.domains.products.actions.save')
    @patch('app.domains.excel.import_excel.models.type')
    @patch('app.domains.products.actions.load_workbook')
    @patch.object(ImportExcel, 'read_cells')
    def test_upload_file_must_be_return_a_product(self, mock_read_cells, load_workbook_mock,
                                                  type_mock, mock_save, mock_product, validate_mock):
        product = [{'name': 'product_name', 'sku': '09876123456', 'shelf_life': '12meses'}]
        file = {'name': 'products.xlsx'}
        file['name'] = tempfile.mkdtemp() + '/products.xlsx'
        file['file'] = (io.BytesIO(bytes(1)), 'products.xlsx')
        load_workbook_mock.return_value = Workbook
        type_mock.return_value = Workbook
        mock_read_cells.return_value = product
        validate_mock.return_value = []
        mock_product.return_value = None
        mock_save.return_value = product

        product_saved = upload_file(file['name'])
        self.assertEqual(product_saved, [product])
        mock_save.called_once_with(product)
        validate_mock.called_once_with(product)
        type_mock.called_once_with(product)
        mock_read_cells(product)


    @patch.object(ValidateCreateProduct, 'check_if_product_of_excel_is_valid')
    @patch('app.domains.products.actions.Product')
    @patch('app.domains.products.actions.save')
    @patch('app.domains.excel.import_excel.models.type')
    @patch('app.domains.products.actions.load_workbook')
    @patch.object(ImportExcel, 'read_cells')
    def test_upload_file_must_be_raise_an_exception(self, mock_read_cells, load_workbook_mock,
                                                  type_mock, mock_save, mock_product, validate_mock):
        product = [{'name': 'product_name', 'sku': '09876123456', 'shelf_life': '12meses'}]
        file = {'name': 'products.xlsx'}
        file['name'] = tempfile.mkdtemp() + '/products.xlsx'
        file['file'] = (io.BytesIO(bytes(1)), 'products.xlsx')
        msg ='error msg'
        load_workbook_mock.return_value = Workbook
        type_mock.return_value = Workbook
        mock_read_cells.return_value = product
        validate_mock.return_value = [msg]
        mock_product.return_value = None
        mock_save.side_effect = ProductExceptions(msg)
        with pytest.raises(ProductExceptions) as exc:
            upload_file(file['name'])
        self.assertEqual(exc.value.code, 406)
        self.assertEqual(exc.value.description, [['error msg']])
        mock_save.called_once_with(product)
        validate_mock.called_once_with(product)
        type_mock.called_once_with(product)
        mock_read_cells(product)



